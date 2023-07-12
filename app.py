import sqlite3
#from jinja2 import Template
from flask import Flask, render_template, request, session,flash, send_file, jsonify
import os
from openpyxl import Workbook
import shutil
import pyperclip
from bancoComputadores import *
from agenda import *

app = Flask(__name__)
#app.secret_key = 'sua_chave_secreta'
app.secret_key = 'segurancati'
logado = []

# rota principal
@app.route('/')
def index():
    directory = 'static/images/Banners'
    # Imagens banner do endomarketing
    # Obtém a lista de arquivos existentes no diretório
    file_list = os.listdir(directory)
    
    print('arquivos',file_list)
    return render_template('index.html',imagens=file_list)

@app.route('/home', methods=['POST'])
def home():
    directory = 'static/images/Banners'
    
    # Obtém a lista de arquivos existentes no diretório
    file_list = os.listdir(directory)
    
    print('arquivos',file_list)
    return render_template('index.html',imagens=file_list)

# Cadastro de Ramais
@app.route('/paginaCadastro',methods=['POST'])
def paginaCadastro():
    return render_template('cadastrar.html')

# Cadastrando Ramal Direto
@app.route('/cadastrarDireto', methods=['POST'])
def cadastrarDireto():
    nome_form = request.form['nome']
    nome_minusculo = nome_form.lower()
    nome = nome_minusculo
    ramal = int(request.form['ramal'])
    setor = request.form['setor']
    # Verificando se existe o ramal
    import sqlite3
    conexao = sqlite3.connect('agenda.db')
    c = conexao.cursor()
    c.execute('SELECT nome, ramal FROM agendaDireta')
    ramais = c.fetchall()
    conexao.close()
    validador = 0
    for i in ramais:
        #print('Ramal cadastrado: '+(i[1])+'Ramal novo:'+ramal)
        if(i[0]==nome):
            retorno = 'Nome já está cadastrado!\n Verifique se é a mesma pessoa ou adicione mais alguma identificação'
            validador = 1
            print("Nome encontrado no banco")
        if(i[1]==ramal):
            retorno = 'Ramal já está cadastrado'
            validador = 1
            print("Ramal encontrado no banco")
        print(i)

    # Verificando se podemos cadastrar o ramal
    if (validador==0):
        conexao = sqlite3.connect('agenda.db')
        c = conexao.cursor()
        c.execute("INSERT INTO agendaDireta (nome, ramal, setor) VALUES (?, ?, ?)", (nome, ramal, setor))

        conexao.commit()
        conexao.close()

        retorno = 'Ramal Direto: '+str(ramal)+' - '+nome+' - '+setor+' \ncadastrado com sucesso!'

        #return f'Obrigado, {nome}, nós entraremos em contato em breve no email {ramal}.'
        #textoRetorno = 'Ramal '+ramal+'-'+nome+' cadastrado com sucesso!'
        #return render_template('cadastrar.html', error=textoRetorno)

    return render_template('cadastrar.html', mensagemRetorno=retorno)

# rota para receber dados do formulário
@app.route('/cadastrar', methods=['POST'])
def cadastrar():
    import sqlite3
    nome_form = request.form['nome']
    nome_minusculo = nome_form.lower()
    nome = nome_minusculo
    ramal = int(request.form['ramal'])
    setor = request.form['setor']
    # Verificando se existe o ramal
    conexao = sqlite3.connect('agenda.db')
    c = conexao.cursor()
    c.execute('SELECT nome, ramal FROM agendaInterna')
    ramais = c.fetchall()
    conexao.close()
    validador = 0
    for i in ramais:
        #print('Ramal cadastrado: '+(i[1])+'Ramal novo:'+ramal)
        if(i[0]==nome):
            retorno = 'Nome já está cadastrado!\n Verifique se é a mesma pessoa ou adicione mais alguma identificação'
            validador = 1
            print("Nome encontrado no banco")
        if(i[1]==ramal):
            retorno = 'Ramal já está cadastrado'
            validador = 1
            print("Ramal encontrado no banco")
        print(i)

    # Verificando se podemos cadastrar o ramal
    if (validador==0):
        conexao = sqlite3.connect('agenda.db')
        c = conexao.cursor()
        c.execute("INSERT INTO agendaInterna (nome, ramal, setor) VALUES (?, ?, ?)", (nome, ramal, setor))

        conexao.commit()
        conexao.close()

        retorno = 'Ramal Interno: '+str(ramal)+' - '+nome+' - '+setor+' \ncadastrado com sucesso!'

        #return f'Obrigado, {nome}, nós entraremos em contato em breve no email {ramal}.'
        #textoRetorno = 'Ramal '+ramal+'-'+nome+' cadastrado com sucesso!'
        #return render_template('cadastrar.html', error=textoRetorno)

    return render_template('cadastrar.html', mensagemRetorno=retorno)
@app.route('/paginaCadastroDireto',methods=['POST'])
def paginaCadastroDireto():
    return render_template('cadastrarDireto.html')

# Editar ----------------------------------------------------------------------------------

# Abrindo a pagina de busca
@app.route('/search_term', methods=['POST'])
def search_term():
    return render_template('search_form.html')

def eh_int(string):
    try:
        int(string)
        return True
    except ValueError:
        return False

# Abrindo a pagina de resultado
@app.route('/search',  methods=['POST'])
def search():
    if request.method == 'POST':
        # Get the search term from the form data
        busca_term = request.form['busca_term']
        nome_minusculo = busca_term.lower()
        nome = nome_minusculo
        if eh_int(busca_term):
            ramal = int(busca_term)
        else:
            ramal = ''
        import sqlite3
        conexao = sqlite3.connect('agenda.db')
        c = conexao.cursor()
        c.execute('SELECT nome, ramal FROM agendaInterna')
        ramais = c.fetchall()
        #conexao.close()
        encontrados = 0
        tipos = []
        listaRamaisI = []
        listaRamaisD = []
        for i in ramais:
            #print('Ramal cadastrado: '+(i[1])+'Ramal novo:'+ramal)
            if(i[0]==nome):
                encontrados = encontrados+1
                tipos.append('interno')
                print("Nome encontrado no banco")
                listaRamaisI.append(i)
                #return render_template('search_select.html')
                #return render_template('search_results.html', search_results=str(i[0]),resultado_ramal=str(i[1]),tipo_ramal='interno')
            if(i[1]==ramal):
                encontrados = encontrados+1
                tipos.append('interno')
                print("Ramal encontrado no banco")
                listaRamaisI.append(i)
                #return render_template('search_select.html')
                #return render_template('search_results.html', search_results=str(i[0]),resultado_ramal=str(i[1]),tipo_ramal='interno')
        c.execute('SELECT nome, ramal FROM agendaDireta')
        ramais = c.fetchall()
        conexao.close()
        for i in ramais:
            #print('Ramal cadastrado: '+(i[1])+'Ramal novo:'+ramal)
            if(i[0]==nome):
                encontrados = encontrados+1
                tipos.append('direto')
                print("Nome encontrado no banco")
                listaRamaisD.append(i)
                #return render_template('search_results.html', search_results=str(i[0]),resultado_ramal=str(i[1]),tipo_ramal='direto')
            if(i[1]==ramal):
                encontrados = encontrados+1
                tipos.append('direto')
                print("Ramal encontrado no banco")
                listaRamaisD.append(i)
                #return render_template('search_results.html', search_results=str(i[0]),resultado_ramal=str(i[1]),tipo_ramal='direto')

    if (encontrados==0):
        return render_template('search_form.html')
    elif (encontrados==1):
        if(tipos[0]=='interno'):
            import sqlite3
            conexao = sqlite3.connect('agenda.db')
            c = conexao.cursor()
            c.execute('SELECT nome, ramal FROM agendaInterna')
            ramais = c.fetchall()
            conexao.close()
            for i in ramais:
                #print('Ramal cadastrado: '+(i[1])+'Ramal novo:'+ramal)
                if(i[0]==nome):
                    print("ENVIANDO")
                    return render_template('search_results.html', search_results=str(i[0]),resultado_ramal=str(i[1]),tipo_ramal='interno')
                if(i[1]==ramal):
                    print("ENVIANDO")
                    return render_template('search_results.html', search_results=str(i[0]),resultado_ramal=str(i[1]),tipo_ramal='interno')
            #return render_template('search_results.html', search_results=str(ramais[]),resultado_ramal=str(ramais[]),tipo_ramal=tipos[0])
        if(tipos[0]=='direto'):
            import sqlite3
            conexao = sqlite3.connect('agenda.db')
            c = conexao.cursor()
            c.execute('SELECT nome, ramal FROM agendaDireta')
            ramais = c.fetchall()
            conexao.close()
            for i in ramais:
                #print('Ramal cadastrado: '+(i[1])+'Ramal novo:'+ramal)
                if(i[0]==nome):
                    print("ENVIANDO")
                    return render_template('search_results.html', search_results=str(i[0]),resultado_ramal=str(i[1]),tipo_ramal='direto')
                if(i[1]==ramal):
                    print("ENVIANDO")
                    return render_template('search_results.html', search_results=str(i[0]),resultado_ramal=str(i[1]),tipo_ramal='direto')
    elif (encontrados>1):
        return render_template('search_select.html', rows=listaRamaisI, rows2=listaRamaisD)
    return render_template('search_select.html')


@app.route('/search_select',  methods=['POST'])
def search_select():
    import sqlite3
    conexao = sqlite3.connect('agenda.db')
    c = conexao.cursor()
    c.execute('SELECT nome, ramal, setor, id FROM agendaInterna')
    ramais = c.fetchall()

    listaRamaisI = []
    listaRamaisD = []
    tipos = []
    for i in ramais:
        tipos.append('interno')
        listaRamaisI.append(i)

    c.execute('SELECT nome, ramal, setor, id FROM agendaDireta')
    ramais = c.fetchall()
    conexao.close()
    for i in ramais:
        tipos.append('direto')
        listaRamaisD.append(i)
    return render_template('search_select.html', rows=listaRamaisI, rows2=listaRamaisD)

@app.route('/buscar_ramal_search_select',  methods=['POST'])
def buscar_ramal_search_select():
    busca = request.form['busca_term']
    rows = select_like(busca, str(busca))
    return render_template('search_select.html', rows=rows)

# Selecionando no banco um ramal parecido com a busca
def select_like(int, str):
    conn = sqlite3.connect('agenda.db')
    c = conn.cursor()
    #cur.execute('SELECT * FROM agendaInterna')
    c.execute("SELECT nome, ramal, setor FROM agendaInterna WHERE ramal LIKE '%' || ? || '%' OR nome LIKE '%' || ? || '%'", (int,str))
    rows = c.fetchall()
    c.execute("SELECT nome, ramal, setor FROM agendaDireta WHERE ramal LIKE '%' || ? || '%' OR nome LIKE '%' || ? || '%'", (int,str))
    rows.extend(c.fetchall())
    conn.close()
    return rows
# Selecionando no banco um ramal por id
def select_id(int,str):
    if str=='interno':
        conn = sqlite3.connect('agenda.db')
        c = conn.cursor()
        #cur.execute('SELECT * FROM agendaInterna')
        c.execute("SELECT nome, ramal, setor FROM agendaInterna WHERE id = ?", (int,))
    elif str=='direto':
        conn = sqlite3.connect('agenda.db')
        c = conn.cursor()
        #cur.execute('SELECT * FROM agendaInterna')
        c.execute("SELECT nome, ramal, setor FROM agendaDireta WHERE id = ?", (int,))
    rows = c.fetchall()
    conn.close()
    return rows

@app.route('/search_results_table',  methods=['POST'])
def search_results_table():
    dados = request.form['editar']
    info = dados.replace(")", "")
    info = info.replace("(", "")
    info = info.replace("'", "")
    info = info.replace(" ", "")
    info = info.split(',')
    print(info)
    data = select_id(info[0],info[1]) # chama a função que busca pelo id e retorna as informações
    infodb = []
    for dados in data[0]:
        infodb.append(dados)
    return render_template('search_results.html',nome=infodb[0],ramal=infodb[1],setor=infodb[2],id=info[0],tipo_ramal=info[1])

@app.route('/search_results',  methods=['POST'])
def search_results():
    return render_template('search_results.html')

@app.route('/salvar_alteracao', methods=['POST'])
def salvar_alteracao():
    tipo_ramal = request.form['tipo_ramal']
    id = request.form['id']
    #print(tipo_ramal)
    if request.form['action'] == 'cadastrar':
        conexao = sqlite3.connect('agenda.db')
        c = conexao.cursor()
        nome = request.form['nome']
        resultado_ramal = request.form['resultado_ramal']
        resultado_setor = request.form['resultado_setor']
        nome = nome.lower()
        ramal = int(resultado_ramal)
        setor = resultado_setor.lower()
        if (tipo_ramal=='interno'):
            c.execute('UPDATE agendaInterna SET nome = ?, ramal = ?, setor = ? WHERE id = ?', (nome, ramal, setor, id))
            conexao.commit()
            conexao.close()
            retorno = 'Ramal Interno: '+str(ramal)+' - '+nome+' - '+setor+' alterado com sucesso!'
            return render_template('admin.html', mensagemRetorno=retorno)
        elif(tipo_ramal=='direto'):
            c.execute('UPDATE agendaDireta SET nome = ?, ramal = ?, setor = ? WHERE ramal = ?', (nome, ramal, setor, ramal))
            conexao.commit()
            conexao.close()
            retorno = 'Ramal Direto: '+str(ramal)+' - '+nome+' - '+setor+' alterado com sucesso!'
            return render_template('admin.html', mensagemRetorno=retorno)
    elif request.form['action'] == 'excluir':
        conexao = sqlite3.connect('agenda.db')
        search_results = request.form['search_results']
        resultado_ramal = request.form['resultado_ramal']
        nome = search_results
        ramal = int(resultado_ramal)
        if(tipo_ramal=='interno'):
            c = conexao.cursor()
            c.execute("DELETE FROM agendaInterna WHERE id=?", (id,))
            conexao.commit()
            conexao.close()

            retorno = 'Ramal Interno: '+str(ramal)+' - '+nome+' excluido com sucesso!'
            return render_template('admin.html', mensagemRetorno=retorno)
        if(tipo_ramal=='direto'):
            c = conexao.cursor()
            c.execute("DELETE FROM agendaDireta WHERE id=?", (id,))
            conexao.commit()
            conexao.close()

            retorno = 'Ramal Direto: '+str(ramal)+' - '+nome+' excluido com sucesso!'
            return render_template('admin.html', mensagemRetorno=retorno)
        return render_template('admin.html', mensagemRetorno='Erro: conflito interno')
@app.route('/abrirAgenda', methods=['POST'])
def abrirAgenda():
    # Importando agenda interna
    conn = sqlite3.connect('agenda.db')
    cur = conn.cursor()
    #cur.execute('SELECT * FROM agendaInterna')
    cur.execute('SELECT nome, ramal, setor FROM agendaInterna ORDER BY nome')
    rows = cur.fetchall()
    # Importando agenda direta
    conexao = sqlite3.connect('agenda.db')
    c = conexao.cursor()
    c.execute('SELECT nome, ramal, setor FROM agendaDireta ORDER BY nome')
    rows2 = c.fetchall()
    #rows.extend(rows2)
    return render_template('table.html', rows=rows, rows2=rows2, ativados=['interno','direto'], filtro='desativado')
@app.route('/abrirFiltro', methods=['POST'])
def abrirFiltro():
    abrir = 'sim'
    return render_template('table.html', ativados=['interno','direto'], filtro='desativado',abrir=abrir)
@app.route('/abrirAgendaFiltrando', methods=['POST'])
def abrirAgendaFiltrando():
    filtros = request.form.getlist('tipo')
    if (filtros==''):
        filtros = ['interno','direto']
    print(filtros)
    rows = []
    rows2 = []
    if (len(filtros)>1):
        if (filtros[0]=='interno') and (filtros[1]=='direto'):
            return abrirAgenda()
        
    for i in filtros:
        if (i=='interno'):
            # Importando agenda interna
            conn = sqlite3.connect('agenda.db')
            cur = conn.cursor()
            #cur.execute('SELECT * FROM agendaInterna')
            cur.execute('SELECT nome, ramal, setor FROM agendaInterna ORDER BY nome')
            rows = cur.fetchall()
        # Importando agenda direta
        if(i=='direto'):
            conexao = sqlite3.connect('agenda.db')
            c = conexao.cursor()
            c.execute('SELECT nome, ramal, setor FROM agendaDireta ORDER BY nome')
            rows2 = c.fetchall()
    rows.extend(rows2)
    return render_template('ramais.html', rows=rows, ativados=filtros, filtro='desativado')

@app.route('/filtrarAgenda', methods=['POST'])
def filtrarAgenda():
    pesquisar = request.form['pesquisar']
    if (pesquisar==''):
        print('Consulta vazia')
    else:
        filtros = pesquisar.split()
        listaFiltros = []
        listaFiltros2 = []
        import sqlite3
        conexao = sqlite3.connect('agenda.db')
        c = conexao.cursor()
        for i in filtros:
            c.execute("SELECT nome, ramal FROM agendaInterna WHERE ramal LIKE '%' || ? || '%' OR nome LIKE '%' || ? || '%'", (i,str(i)))
            listaFiltros.append(c.fetchall())
        for i in filtros:
            c.execute("SELECT nome, ramal FROM agendaDireta WHERE ramal LIKE '%' || ? || '%' OR nome LIKE '%' || ? || '%'", (i,str(i)))
            listaFiltros2.append(c.fetchall())
        conexao.close()
        ramais = listaFiltros
        ramais2 = listaFiltros2
        print(listaFiltros)
        return render_template('table.html', rows=ramais, rows2=ramais2,filtros=filtros,filtro='ativado')
        #return render_template('table.html', rows=rows, rows2=rows2)
    return abrirAgenda()

@app.route('/excluirFiltro', methods=['POST'])
def excluirFiltro():
    # O botão vai retornar uma string ['nomeExemplo',('nomeExemplo2','nomeExemplo3')]
    # A string vai passar por todos os filtros para remover os caracteres e separar em uma lista
    filtros = request.form['filtro']
    listaFiltros = filtros.replace("(", "")
    listaFiltros = listaFiltros.replace(")", "")
    listaFiltros = listaFiltros.replace("[", "")
    listaFiltros = listaFiltros.replace("]", "")
    listaFiltros = listaFiltros.replace("'", "")
    listaFiltros = listaFiltros.replace(" ", "")
    listaFiltros = listaFiltros.split(",")
    # A primeira posição da lista é o filtro que será excluido
    filtroExcluido = listaFiltros[0]
    del listaFiltros[0]
    # Agora vamos percorrer a lista buscando a posição do filtro para ser excluido
    tamanhoLista = len(listaFiltros)
    i=0
    while (i<tamanhoLista):
        if (listaFiltros[i]==filtroExcluido):
            posicaoFiltroExcluido = i
        i = i+1
    # Deletando o filtro dentro da lista
    del listaFiltros[posicaoFiltroExcluido]
    for i in listaFiltros:
        print('Restante da lista:',i)
    print('tamanho da lista',len(listaFiltros))
    # Filtrando novamente e retornando
    if ((len(listaFiltros))>0):
        print('entrou')
        filtros = listaFiltros
        ramaisFiltrados = []
        ramaisFiltrados2 = []
        import sqlite3
        conexao = sqlite3.connect('agenda.db')
        c = conexao.cursor()
        for i in filtros:
            c.execute("SELECT nome, ramal FROM agendaInterna WHERE ramal LIKE '%' || ? || '%' OR nome LIKE '%' || ? || '%'", (i,i))
            ramaisFiltrados.append(c.fetchall())
        for i in filtros:
            c.execute("SELECT nome, ramal FROM agendaDireta WHERE ramal LIKE '%' || ? || '%' OR nome LIKE '%' || ? || '%'", (i,str(i)))
            ramaisFiltrados2.append(c.fetchall())
        conexao.close()
        ramais = ramaisFiltrados
        ramais2 = ramaisFiltrados2
        return render_template('table.html', rows=ramais, rows2=ramais2,filtros=filtros,filtro='ativado')
    else:
        print('entrou')
        return abrirAgenda()

@app.route('/login', methods=['POST'])
def login():
    print('login')
    if 'logged_in' in session:
        return render_template('admin.html',nome=session['usuario'],acesso=session['nivel_acesso'])
    else:
        return render_template('login.html')

@app.route('/admin_retorno', methods=['POST'])
def admin_retorno(mensagem, submensagem):
    print('admin_retorno')
    flash(mensagem+'<br>')
    flash(submensagem)
    if 'logged_in' in session:
        return render_template('admin.html',nome=session['usuario'],acesso=session['nivel_acesso'])
    else:
        return render_template('login.html')

@app.route('/validarLogin', methods=['POST'])
def validarLogin():
    usuario = request.form['usuario']
    senha = request.form['senha']
    # Importando agenda interna
    conn = sqlite3.connect('users.db')
    cur = conn.cursor()
    #cur.execute('SELECT * FROM agendaInterna')
    cur.execute('SELECT usuario, senha, acesso, id FROM usuarios ORDER BY usuario')
    lista = cur.fetchall()
    for i in lista:
        if i[0]==usuario:
            print('Usuario encontrado')
            if (i[1]==senha):
                print('Senha validada, acesso liberado!')
                print('ID: ',i[3])
                print('Nivel de acesso: ',i[2])
                logado.append(i)
                print('logado',logado)
                session['logged_in'] = True
                session['nivel_acesso'] = i[2]
                session['usuario'] = usuario
                return render_template('admin.html', nome=usuario, acesso=i[2], id=i[3])
    session.pop('logged_in', None)
    return render_template('login.html',mensagemRetorno='Usuario ou senha incorreto')

@app.route('/admin', methods=['POST'])
def admin():
    return render_template('admin.html')

@app.route('/logout', methods=['POST'])
def logout():
    session.pop('logged_in', None)
    return home()

@app.route('/editarBanner', methods=['POST'])
def editarBanner():
    return render_template('editarBanner.html')

# Adicionar e excluir imagens do diretorio

UPLOAD_FOLDER = 'static/images/Banners'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
#app.config['ALLOWED_EXTENSIONS'] = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'}


@app.route('/uploadDrop', methods=['POST'])
def uploadDrop():
    file = request.files['file']
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))
    retorno = 'Arquivo {} enviado com sucesso!'.format(file.filename)
    return render_template('editarBanner.html',mensagemRetorno=retorno)


@app.route('/upload', methods=['POST'])
def upload():
    # Verifica se o arquivo está presente no request
    if 'file' not in request.files:
        retorno = 'Nenhum arquivo enviado'
        return render_template('editarBanner.html',mensagemRetorno=retorno)

    file = request.files['file']

    # Verifica se o arquivo possui um nome
    if file.filename == '':
        retorno = 'Nome do arquivo não encontrado'
        return render_template('editarBanner.html',mensagemRetorno=retorno)

    # Salva o arquivo no diretório de upload
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))


    retorno = 'Arquivo {} enviado com sucesso!'.format(file.filename)
    return render_template('editarBanner.html',mensagemRetorno=retorno)

@app.route('/excluirImagem', methods=['POST'])
def excluirImagem():
    botaoExcluir = request.form['botaoExcluir']
    print(botaoExcluir)
    # Caminho completo para o arquivo que será excluído
    nomeArquivo = 'static/images/Banners/'+botaoExcluir
    file_path = nomeArquivo
    
    # Verifica se o arquivo existe
    if os.path.exists(file_path):
        # Exclui o arquivo
        os.remove(file_path)
        retorno = 'Arquivo excluído com sucesso!'
        return render_template('editarBanner.html',mensagemRetorno=retorno)
    else:
        retorno = 'Arquivo não encontrado.'
        return render_template('editarBanner.html',mensagemRetorno=retorno)

@app.route('/listarDiretorio', methods=['POST'])
def listarDiretorio():
    # Diretório onde estão os arquivos
    directory = 'static/images/Banners'
    
    # Obtém a lista de arquivos existentes no diretório
    file_list = os.listdir(directory)
    
    print('arquivos',file_list)

    return render_template('editarBanner.html', files=file_list)
@app.route('/renomearImagem', methods=['POST'])
def renomearImagem():
    # Diretório onde estão as imagens
    directory = 'static/images/Banners'

    nomeNovo = request.form['nome_novo']
    nomeAntigo = request.form['nome_antigo']
    # Nome antigo do arquivo
    old_name = nomeAntigo
    
    # Novo nome desejado para o arquivo
    new_name = nomeNovo
    
    # Caminho completo para o arquivo antigo
    old_path = os.path.join(directory, old_name)
    
    # Caminho completo para o arquivo com o novo nome
    new_path = os.path.join(directory, new_name)
    
    # Verifica se o arquivo antigo existe
    if os.path.exists(old_path):
        # Renomeia o arquivo
        os.rename(old_path, new_path)
        retorno = 'Arquivo renomeado com sucesso!'
        return render_template('editarBanner.html',mensagemRetorno=retorno)
    else:
        retorno = 'Arquivo não encontrado.'
        return render_template('editarBanner.html',mensagemRetorno=retorno)
@app.route('/ti', methods=['POST'])
def ti():
    return render_template('TI.html')

@app.route('/download')
def download_file():
    # Caminho para o arquivo que você deseja disponibilizar para download
    path_to_file = '/static/download/teste.txt'

    return "<a href='{{ url_for('get-file') }}'>Download</a>"

@app.route('/get-file')
def get_file():
    # Caminho para o arquivo que você deseja disponibilizar para download
    path_to_file = '/static/download/teste.txt'

    # Envia o arquivo para o cliente
    return send_file(path_to_file, as_attachment=True)

@app.route('/copiar_link', methods=['POST'])
def copiar_link():
    link = request.form['link']
    pyperclip.copy(link)
    flash('Link copiado para a Área de Transferência (CTRL+V)\n'+'\n'+link)
    return render_template('TI.html')

@app.route('/tabela_ip', methods=['POST'])
def tabela_ip():
    import sqlite3
    conexao = sqlite3.connect('tabelaIP.db')
    c = conexao.cursor()
    #c.execute('SELECT ip, mac, usuario, host, setor, dhcp FROM tabelaIP')
    c.execute('SELECT * FROM tabelaIP ORDER BY ip')
    tabelaIP = c.fetchall()
    conexao.close()
    #print(tabelaIP)
    return render_template('tabelaip.html', tabela=tabelaIP,resultados=len(tabelaIP))
def testar_ip(valor):
    """if isinstance(valor, int):
        print("O valor é um inteiro.")
    elif isinstance(valor, str):
        print("O valor é uma string.")
    else:
        print("O valor não é um inteiro nem uma string.")"""
    try:
        valor_int = int(valor)
        print("A conversão foi bem-sucedida. O valor inteiro é:", valor_int)
        return valor_int
    except ValueError:
        print("A conversão falhou. Vamos testar se é um IP valido.")
    if (len(valor)>3):
        if (valor[:3]=='192'): # Começa com 192
            print('É um IP') #192.168.20.1
        if valor[:3]=='20.':
            print('É uma pesquisa na faixa 20 com os numeros finais')
            ip = valor[3:]
            return ip
        if (len(valor)>8): # Testando se vai até a 192.168.20.
            if (valor[8:10]=='20'): # Testando se é faixa 20
                print('É um ip faixa 20')
                if (len(valor)>11): # Pegando apenas o valor do IP
                    ip = valor[11:]
                    print('IP final: ',ip)
                    return ip
        return None
def testar_mac(valor):
    if (valor[2:3]==':'):
        print("É um MAC separado por :")
        return valor
    if (valor[2:3]=='-'):
        print("É um MAC separado por -")
        mac = valor.replace("-", ":")
        return mac
    return None
def testar_pesquisa(str):
    print('Buscando correspondencia do usuario, setor e host')
    import sqlite3
    conexao = sqlite3.connect('tabelaIP.db')
    c = conexao.cursor()
    c.execute("SELECT * FROM tabelaIP WHERE usuario LIKE '%' || ? || '%' OR setor LIKE '%' || ? || '%' OR host LIKE '%' || ? || '%'", (str,str,str))
    tabelaIP = c.fetchall()
    conexao.close()
    if (len(tabelaIP)!=0):
        return tabelaIP
    return None
@app.route('/filtrar_tabela_ip', methods=['POST'])
def filtrar_tabela_ip():
    dados = request.form['pesquisar']
    if (dados==''):
        return tabela_ip()
    else:
        ip = testar_ip(dados) # reduzindo ip para apenas os numeros finais
        mac = testar_mac(dados) # formatando a pesquisa MAC
        usuario_setor_host = testar_pesquisa(dados) # Buscando por correspondencias de usuario, host e setor e inserindo na lista
        if (ip==None) and (mac==None) and (usuario_setor_host==None):
            return render_template('tabelaip.html',vazio='vazio')
        import sqlite3
        conexao = sqlite3.connect('tabelaIP.db')
        c = conexao.cursor()
        #c.execute('SELECT ip, mac, usuario, host, setor, dhcp FROM tabelaIP')
        #c.execute('SELECT * FROM tabelaIP WHERE ip = ? OR mac = ? OR usuario = ? OR setor = ? ORDER BY ip', (ip,mac,dados,dados)
        c.execute('SELECT * FROM tabelaIP WHERE ip = ? OR mac = ? ORDER BY ip', (ip,mac))
        tabela_IP_MAC = c.fetchall()
        tabela = []
        tabela.extend(tabela_IP_MAC)
        # Se existem usuarios, host ou setores compativeis com a pesquisa serão adicionados a tabela
        if (usuario_setor_host!=None): 
            tabela.extend(usuario_setor_host)
        conexao.close()
        if (len(tabela) == 1):
            if(tabela[0]==[]):
                print('vazio')
                return render_template('tabelaip.html',vazio='vazio')
        print(tabela)
        return render_template('tabelaip.html', tabela=tabela,resultados=len(tabela))

@app.route('/filtrar_ip_livre_ocupado', methods=['POST'])
def filtrar_ip_livre_ocupado():
    filtros = request.form.getlist('estado')
    if (filtros==''):
        filtros = ['usados','livres']
    print(filtros)
    tabela = []
    tabela2 = []
    if (len(filtros)>1):
        if (filtros[0]=='usados') and (filtros[1]=='livres'):
            return tabela_ip()
        
    for i in filtros:
        if (i=='usados'):
            # Importando agenda interna
            conn = sqlite3.connect('tabelaIP.db')
            cur = conn.cursor()
            cur.execute('SELECT * FROM tabelaIP WHERE mac IS NOT NULL OR usuario IS NOT NULL ORDER BY ip')
            tabela = cur.fetchall()
        # Importando agenda direta
        if(i=='livres'):
            usuario = None
            mac = None
            conexao = sqlite3.connect('tabelaIP.db')
            c = conexao.cursor()
            c.execute('SELECT * FROM tabelaIP WHERE usuario IS NULL OR mac IS NULL ORDER BY ip')
            #c.execute('SELECT * FROM tabelaIP')
            tabela2 = c.fetchall()
    tabela.extend(tabela2)
    print(tabela)
    return render_template('tabelaIP.html', tabela=tabela, ativados=filtros,resultados=len(tabela))


@app.route('/editar_ip', methods=['POST'])
def editar_ip():
    dados = request.form['editar']
    import sqlite3
    conexao = sqlite3.connect('tabelaIP.db')
    c = conexao.cursor()
    #c.execute('SELECT ip, mac, usuario, host, setor, dhcp FROM tabelaIP')
    #c.execute('SELECT * FROM tabelaIP ORDER BY ip')
    c.execute("SELECT * FROM tabelaIP WHERE id = ?", (dados,))
    tabelaIP = c.fetchall()
    return render_template('editar_ip.html',dados=tabelaIP)

@app.route('/salvar_edicao_ip', methods=['POST'])
def salvar_edicao_ip():
    id_ip = request.form['id_ip']
    ip = request.form['ip']
    ip = ip[11:] # Removendo o "192.168.20." para salvar apenas o ip final
    mac = request.form['mac']
    usuario = request.form['usuario']
    host = request.form['host']
    setor = request.form['setor']
    dhcp = request.form['dhcp']
    if (request.form['mac']=='')or(request.form['mac']=='None'): # Conferindo se o MAC veio em branco ou NoneType
        mac = None
    if (request.form['usuario']=='')or(request.form['usuario']=='None'): # Conferindo se o usuario veio em branco ou NoneType
        usuario = None
    if (request.form['host']=='')or(request.form['host']=='None'): # Conferindo se o host veio em branco ou NoneType
        host = None
    if (request.form['setor']=='')or(request.form['setor']=='None'): # Conferindo se o setor veio em branco ou NoneType
        setor = None
    if (request.form['dhcp']=='')or(request.form['dhcp']=='None'): # Conferindo se o dhcp veio em branco ou NoneType
        dhcp = None
    conexao = sqlite3.connect('tabelaIP.db')
    c = conexao.cursor()
    c.execute('UPDATE tabelaIP SET ip = ?, mac = ?, usuario = ?, host = ?, setor = ?, dhcp = ? WHERE id = ?', (ip, mac, usuario, host, setor, dhcp, id_ip))
    conexao.commit()
    conexao.close()

    #return render_template('login.html',mensagemRetorno='Alteração concluida')
    mensagem = 'IP alterado com sucesso!'
    submensagem = '192.168.20.'+ip
    return admin_retorno(mensagem, submensagem)

def buscar_login():
    # Importando agenda interna
    conn = sqlite3.connect('users.db')
    cur = conn.cursor()
    #cur.execute('SELECT * FROM agendaInterna')
    cur.execute('SELECT usuario, senha, acesso, id FROM usuarios ORDER BY usuario')
    return cur.fetchall()

@app.route('/editar_logins', methods=['POST'])
def editar_logins():
    lista = buscar_login()
    return render_template('editar_logins.html',users=lista)

@app.route('/selecionar_logins', methods=['POST'])
def selecionar_logins():
    user = request.form['user']
    lista = buscar_login()
    for i in lista:
        if (user==i[0]):
            return render_template('editar_logins.html',user_info=i)
        print(i)
    return render_template('editar_logins.html')

@app.route('/salvar_edicao_logins', methods=['POST'])
def salvar_edicao_logins():
    id = request.form['id']
    usuario = request.form['usuario']
    senha = request.form['senha']
    acesso = request.form['acesso']
    # Importando
    conn = sqlite3.connect('users.db')
    cur = conn.cursor()
    #cur.execute('SELECT * FROM agendaInterna')
    #cur.execute('SELECT usuario, senha, acesso, id FROM usuarios ORDER BY usuario')
    cur.execute('UPDATE usuarios SET usuario = ?, senha = ?, acesso = ? WHERE id = ?', (usuario, senha, acesso, id))
    info = cur.fetchall()
    print(info)
    conn.commit()
    conn.close()

    return login()

@app.route('/criar_novo_login', methods=['POST'])
def criar_novo_login():
    user_new = 'novo'
    return render_template('editar_logins.html', user_new=user_new)

@app.route('/salvar_novo_login', methods=['POST'])
def salvar_novo_login():
    usuario = request.form['usuario']
    senha = request.form['senha']
    acesso = request.form['acesso']

    conn = sqlite3.connect('users.db')
    cur = conn.cursor()
    cur.execute("INSERT INTO usuarios (usuario, senha, acesso) VALUES (?, ?, ?)", (usuario, senha, acesso))
    conn.commit()
    conn.close()

    return login()

@app.route('/excluir_login', methods=['POST'])
def excluir_login():
    print('entrou no excluir')
    id = request.form['id']
    print(id)
    # Importando banco
    conn = sqlite3.connect('users.db')
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE id=?", (id,))
    conn.commit()
    conn.close()

    mensagem = "Usuário excluido com sucesso!"
    submensagem = ''
    return admin_retorno(mensagem, submensagem)

@app.route('/usuarios_wts', methods=['POST'])
def usuarios_wts():
    import sqlite3
    conexao = sqlite3.connect('tabelaWts.db')
    c = conexao.cursor()
    #c.execute('SELECT ip, mac, usuario, host, setor, dhcp FROM tabelaIP')
    c.execute('SELECT * FROM tabelaWts ORDER BY id AND usuario')
    tabelaWts = c.fetchall()
    conexao.close()
    #print(tabelaIP)
    return render_template('usuarios_wts.html', tabela=tabelaWts,resultados=len(tabelaWts))

@app.route('/editar_usuarios_wts', methods=['POST'])
def editar_usuarios_wts():
    dados = request.form['editar']
    import sqlite3
    conexao = sqlite3.connect('tabelaWts.db')
    c = conexao.cursor()
    c.execute('SELECT * FROM tabelaWts WHERE id = ?', (dados,))
    tabelaWts = c.fetchall()
    conexao.close()
    return render_template('editar_wts.html', dados=tabelaWts)

@app.route('/salvar_edicao_usuarios_wts', methods=['POST'])
def salvar_edicao_usuarios_wts():
    id = request.form['id']
    usuario = request.form['usuario']
    senha = request.form['senha']
    setor = request.form['setor']
    ip = request.form['ip']
    ip = ip[11:]

    # Importando banco de dados
    conn = sqlite3.connect('tabelaWts.db')
    cur = conn.cursor()
    cur.execute('UPDATE tabelaWts SET usuario = ?, senha = ?, setor = ?, ip = ? WHERE id = ?', (usuario, senha, setor, ip, id))
    info = cur.fetchall()
    print(info)
    conn.commit()
    conn.close()

    mensagem = 'Usuário alterado com sucesso!'
    submensagem = usuario
    return admin_retorno(mensagem, submensagem)

@app.route('/usuarios_wts_ip', methods=['POST']) # Buscar qual a localização / pc está usando o usuario SIGGI
def usuarios_wts_ip():
    id_usuario = request.form['id_usuario']
    dados = request.form['id_ip']
    import sqlite3
    conexao = sqlite3.connect('tabelaIP.db')
    c = conexao.cursor()
    c.execute('SELECT * FROM tabelaIP WHERE ip = ? ORDER BY ip', (dados,))
    tabelaIp = c.fetchall()

    conexao = sqlite3.connect('tabelaWts.db')
    c = conexao.cursor()
    #c.execute('SELECT ip, mac, usuario, host, setor, dhcp FROM tabelaIP')
    c.execute('SELECT * FROM tabelaWts WHERE id = ? ORDER BY id AND usuario',(id_usuario,))
    tabelaWts = c.fetchall()
    conexao.close()

    if len(tabelaIp)==0:
        print('Nenhum vinculo')
        return render_template('vincular_wts_ip.html',tabelaWts=tabelaWts)
    else:
        return render_template('vincular_wts_ip.html',tabelaIp=tabelaIp,tabelaWts=tabelaWts)

@app.route('/exportar_planilhas', methods=['POST']) # Exportar os dados do banco como planilhas para realizar o download
def exportar_planilhas():
    # criando uma lista de mensagem de retorno
    mensagens = []
    # definindo uma submensagem caso não aconteça nenhum erro durante o processo
    submensagem = 'Para realizar o download clique em Utilidades e escolha a planilha que deseja'
    # Testando erros na primeira exportação
    try:
        ''' Exportando Agenda '''
        # Conexão com banco
        Interna = "SELECT * FROM agendaInterna"
        conexao = sqlite3.connect('agenda.db')
        cursor = conexao.cursor()
        cursor.execute(Interna)
        agendaInterna = cursor.fetchall()

        Direta = "SELECT * FROM agendaDireta"
        conexao = sqlite3.connect('agenda.db')
        cursor = conexao.cursor()
        cursor.execute(Direta)
        agendaDireta = cursor.fetchall()

        # Criando planilha
        book = Workbook()
        sheet = book.active
        nome_xlsx = f"static/download/Ramais.xlsx"

        sheet['A1'] = 'Ramais Internos'
        sheet['C1'] = 'Ramais Diretos'
        sheet['A2'] = 'Ramal'
        sheet['B2'] = 'Nome'
        sheet['C2'] = 'Ramal'
        sheet['D2'] = 'Nome'
        cont = 3
        for i in agendaInterna:
            sheet['A'+str(cont)] = int(i[2])
            sheet['B'+str(cont)] = str(i[1])
            cont = cont + 1
        cont = 3
        for i in agendaDireta:
            sheet['C'+str(cont)] = int(i[2])
            sheet['D'+str(cont)] = str(i[1])
            cont = cont + 1
        book.save(nome_xlsx)
        mensagem = 'Planilha de Ramais exportada com sucesso!'
    except Exception as e:
        erro = f"{str(e)}"
        resumo = erro
        if erro[:47]=="[Errno 13] Permission denied: 'static/download/":
            mensagem = 'Erro ao salvar Planilha de Ramais!<br>Verifique se não está com a planilha original aberta.<br>'
            submensagem = 'Codigo: '+resumo
    mensagens.append(mensagem)
        
    ''' ----------------------- Exportando IPs --------------------------- '''
    
    # Importando banco de dados
    #import sqlite3
    conexao = sqlite3.connect('tabelaIP.db')
    c = conexao.cursor()
    c.execute('SELECT * FROM tabelaIP ORDER BY ip')
    tabelaIp = c.fetchall()
    try:
        # Criando planilha
        book = Workbook()
        sheet = book.active
        nome_xlsx = f"static/download/IPs.xlsx"

        sheet['A1'] = 'IP'
        sheet['B1'] = 'MAC'
        sheet['C1'] = 'Usuarios'
        sheet['D1'] = 'Host'
        sheet['E1'] = 'Setor'
        sheet['F1'] = 'DHCP'

        cont = 2 # define a contagem inicial
        for i in tabelaIp:
            sheet['A'+str(cont)] = '192.168.20.'+str(i[1]) # insere o valor na planilha
            sheet['B'+str(cont)] = i[2]
            sheet['C'+str(cont)] = i[3]
            sheet['D'+str(cont)] = i[4]
            sheet['E'+str(cont)] = i[5]
            sheet['F'+str(cont)] = i[6]
            cont = cont + 1

        book.save(nome_xlsx)

        mensagem = 'Planilha de IPs exportada com sucesso!'

    except Exception as e:
        erro = f"{str(e)}"
        resumo = erro
        if erro[:47]=="[Errno 13] Permission denied: 'static/download/":
            mensagem = 'Erro ao salvar Planilha de IPs!<br>Verifique se não está com a planilha original aberta.<br>'
            submensagem = 'Codigo: '+resumo
    mensagens.append(mensagem)

    ''' ----------------------- Exportando usuarios WTS --------------------------- '''
    # Importando banco de dados
    #import sqlite3
    conexao = sqlite3.connect('tabelaWts.db')
    c = conexao.cursor()
    c.execute('SELECT * FROM tabelaWts ORDER BY id')
    tabelaWts = c.fetchall()
    try:
        # Criando planilha
        book = Workbook()
        sheet = book.active
        nome_xlsx = f"static/download/Usuarios WTS.xlsx"

        sheet['A1'] = 'Usuario'
        sheet['B1'] = 'Senha'
        sheet['C1'] = 'Setor'
        sheet['D1'] = 'IP'

        cont = 2 # define a contagem inicial
        for i in tabelaWts:
            sheet['A'+str(cont)] = str(i[1]) # insere o valor na planilha
            sheet['B'+str(cont)] = i[2]
            sheet['C'+str(cont)] = i[3]
            if i[4]=='None' or i[4]==None:
                sheet['D'+str(cont)] = '-'
            else:
                sheet['D'+str(cont)] = '192.168.20.'+str(i[4])
            cont = cont + 1

        book.save(nome_xlsx)
        mensagem = 'Planilha Usuarios WTS exportada com sucesso!'

    except Exception as e:
        erro = f"{str(e)}"
        resumo = erro
        if erro[:47]=="[Errno 13] Permission denied: 'static/download/":
            mensagem = 'Erro ao salvar Planilha Usuario WTS!<br>Verifique se não está com a planilha original aberta.<br>'
            submensagem = 'Codigo: '+resumo
    mensagens.append(mensagem)

    for mensagem in mensagens:
        flash(mensagem+'<br>')

    flash(submensagem)

    return login()

def fazer_copia_banco(str):
    # Caminho do banco de dados original
    nome_banco = str
    caminho_origem = os.path.join(app.root_path, nome_banco+'.db')

    # Caminho da pasta de destino para a cópia
    pasta_destino = os.path.join(app.root_path)
    pasta_destino = pasta_destino+'/static/download/'
    print('Caminho: ',pasta_destino)

    # Verificar se a pasta de destino existe, caso contrário, criar
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)

    # Nome do arquivo da cópia do banco de dados
    nome_arquivo_copia = nome_banco+'_backup.db'

    # Caminho completo para a cópia do banco de dados
    caminho_copia = os.path.join(pasta_destino, nome_arquivo_copia)

    # Copiar o arquivo do banco de dados para a pasta de destino
    shutil.copy2(caminho_origem, caminho_copia)

    mensagem = 'Cópia do banco de dados '+ nome_banco +' criada com sucesso!'
    submensagem = 'Iniciando download'
    messages = []
    messages.append(mensagem)
    messages.append(submensagem)
    return messages

@app.route('/exportar_banco_agenda',methods=['POST'])
def exportar_banco_agenda():
    nome_banco = request.form['nome_banco']
    # Clonando para o diretorio /static/download/
    messages = fazer_copia_banco(nome_banco)
    # Caminho para o arquivo do banco de dados
    pasta_destino = os.path.join(app.root_path)
    pasta_destino = pasta_destino+'/static/download/'+nome_banco+'_backup.db'
    db_path = pasta_destino
    # Disparar o download do arquivo
    return send_file(db_path, as_attachment=True)
    
    #return render_template('TI.html')

""" Equipamentos """

@app.route('/listar_equipamentos',methods=['POST'])
def listar_equipamentos():
    lista = list(ListarEquipamentos())
    #print(lista)
    return render_template('equipamentos.html',equipamentos=lista[0])

@app.route('/filtrar_equipamentos',methods=['POST'])
def filtrar_equipamentos():
    filtrar = request.form['filtrar-tabela']
    equipamentos = list(FiltrarEquipamentos(filtrar))
    return render_template('equipamentos.html',equipamentos=equipamentos[0])

@app.route('/abrir_menu_lateral',methods=['POST'])
def abrir_menu_lateral():
    return render_template('equipamentos.html',menu_lateral='ativo')

@app.route('/copiar_licenca',methods=['POST'])
def copiar_licenca():
    id_licenca = request.form['id_licenca']
    licenca = select_licenca_id(id_licenca)
    pyperclip.copy(licenca[2])
    #return listar_equipamentos()
    return render_template('equipamentos.html',notificacao='Licença copiada para a Área de Transferência (Ctrl+V)')

@app.route('/listar_licenca',methods=['POST'])
def listar_licenca():
    id_licenca = request.form['id_licenca']
    licenca = select_licenca_id(id_licenca)
    print(licenca[1])
    if (licenca[1]=='Windows 7 Professional x64'):
        versao = 'windows-7'
    elif (licenca[1]=='Windows 10 Pro x64'):
        versao = 'windows-10'
    if (licenca[1]=='Windows 11 Pro x64'):
        versao = 'windows-11'

    id_equipamento = request.form['id_equipamento']
    equipamento = select_equipamento_id(id_equipamento)

    flash(licenca)
    return render_template('equipamentos.html',licenca=licenca, info_equipamento=equipamento, versao=versao)

@app.route('/editar_equipamento',methods=['POST'])
def editar_equipamento():
    id_equipamento = request.form['id_equipamento']
    #print(id_equipamento)
    equipamento = select_equipamento_id(id_equipamento)
    licencas = ListarLicencas()
    #flash(equipamento)
    return render_template('equipamentos.html',equipamento=equipamento, licencas=licencas)

@app.route('/salvar_alteracoes_equipamento',methods=['GET','POST'])
def salvar_alteracoes_equipamento():
    if(request.form['acao']=='salvar'):
        id = request.form['id']
        usuario = request.form['usuario']
        tipo_equipamento = request.form.get('tipo_equipamento')
        ip = request.form['ip']
        licencas = request.form['licencas']
        print(tipo_equipamento)
        if (tipo_equipamento== None ):
            print('Tipo zerado')
            dados_antigos = select_equipamento_id(id)
            tipo_equipamento = dados_antigos[2]
        print(usuario,tipo_equipamento,ip,licencas,id)
        salvar_alteracoes(usuario, tipo_equipamento, ip, licencas, id)
        return render_template('equipamentos.html',notificacao='Operação realizada com sucesso!')
    elif(request.form['acao']=='excluir'):
        excluir_equipamento(request.form['id'])
        return render_template('equipamentos.html',notificacao='Equipamento excluído')
    else:
        return render_template('equipamentos.html',notificacao='Operação cancelada')

""" --------------------------------- Agenda nova ------------------------------------------------- """
# ================================================================================================== #

@app.route('/agenda',methods=['POST'])
def agenda():
    interna = select_interna()
    direta = select_direta()
    if 'logged_in' in session:
        return render_template('agenda.html',nome_login=session['usuario'], interna=interna,direta=direta)
    else:
        return render_template('agenda.html')

if __name__ == '__main__':
    #app.run(host='192.168.20.125')
    app.run(debug=True)